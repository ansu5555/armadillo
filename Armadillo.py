import os.path
import zipfile

from datetime import datetime
from fpdf import FPDF
from logbook import Logger, FileHandler
from openpyxl import *
from ruamel.yaml import YAML
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.ui import WebDriverWait

__driver = lambda: None
__pomdict = {}
__logwriter = lambda: None
__reportwriter = lambda: None
__resultpath = ''
__datapath = ''
__libpath = ''
__starttime = ''
__datahandler = []


# def readxl(sheetname):
#     wb = load_workbook(sheetname)
#     ws = wb['Sheet1']
#     rowcount = len(ws["A"])
#     colcount = len(ws[1])
#     funclist = []
#     for j in range(1, rowcount + 1, 2):
#         funcname = str(ws.cell(row=j, column=1).value)
#         arg = ''
#         for i in range(2, colcount + 1):
#             val = str(ws.cell(row=j + 1, column=i).value)
#             if val == 'None':
#                 break
#             if i == 2:
#                 arg = '"' + val + '"'
#             else:
#                 arg += ',"' + val + '"'
#         func = funcname + '(' + arg + ')'
#         funclist.append(func)
#     return funclist


# Functions for direct calls    ----------------------------------------------------------------------------------------


def dataprovider(testname, variable):
    try:
        value = datahandler[testname][variable]
        return value
    except KeyError:
        reportlogger(False, variable + ' does not exist in data file')


def datakeeper(testname, variable, value):
    try:
        datahandler[testname][variable] = value
    except KeyError:
        datahandler[testname].insert(1, variable, value, )
    finally:
        outfile = open(datapath + '/data.yml', 'w')
        yml = YAML()
        yml.dump(datahandler, outfile)


def reportlogger(result, decription, screenshot=None):
    dtstamp = str(datetime.now().strftime("%d-%m-%Y %H:%M:%S"))
    if result:
        reportwriter.set_text_color(76, 175, 80)
        reportwriter.set_font('ZapfDingbats', '', 14)
        reportwriter.cell(14, 7, '4', '', 0, 'C')
        reportwriter.set_text_color(33, 33, 33)
        reportwriter.set_font('Arial', '', 10)
        reportwriter.cell(120, 7, decription, '', 0)
        reportwriter.cell(20, 7, dtstamp, '', 1)
        logwriter.notice('Step passed: ' + decription)
    else:
        reportwriter.set_text_color(244, 67, 54)
        reportwriter.set_font('ZapfDingbats', '', 14)
        reportwriter.cell(14, 7, '6', '', 0, 'C')
        reportwriter.set_text_color(33, 33, 33)
        reportwriter.set_font('Arial', '', 10)
        reportwriter.cell(120, 7, 'OK leave it why to think   ', '', 0)
        reportwriter.cell(20, 7, dtstamp, '', 1)
        logwriter.error('Step failed: ' + decription)


# class for handling execution  ----------------------------------------------------------------------------------------


class ArmadilloRunner:
    def navigate(self, url):
        driver.get(url)
        print(id(driver))
        reportlogger(True, 'Launched ' + url + ' on ' + driver.capabilities['browserName'])

    def wait(self, seconds):
        driver.implicitly_wait(seconds)
        reportlogger(True, 'Wait for ' + seconds + ' seconds ')

    def getelement(self, pagename, elmtname, wait=20):
        key = pagename + '.' + elmtname
        elmtdtl = pomdict[key].split('|')
        if elmtdtl[0] == 'ID':
            logwriter.info('Get element ' + elmtdtl[1] + ' of type ' + elmtdtl[0] + ' in page ' + pagename)
            elmt = WebDriverWait(driver, wait).until(
                expected_conditions.presence_of_element_located((By.ID, elmtdtl[1])))
        elif elmtdtl[0] == 'XPATH':
            logwriter.info('Get element ' + elmtdtl[1] + ' of type ' + elmtdtl[0] + ' in page ' + pagename)
            elmt = WebDriverWait(driver, wait).until(
                expected_conditions.presence_of_element_located((By.XPATH, elmtdtl[1])))
        elif elmtdtl[0] == 'CLASS':
            logwriter.info('Get element ' + elmtdtl[1] + ' of type ' + elmtdtl[0] + ' in page ' + pagename)
            elmt = WebDriverWait(driver, wait).until(
                expected_conditions.presence_of_element_located((By.CLASS_NAME, elmtdtl[1])))
        else:
            logwriter.error('Element ' + elmtdtl[1] + ' not found of type ' + elmtdtl[0] + ' in page ' + pagename)
            print('wrong type cannot be identified')
        return elmt

# Constructor and Destructor    ----------------------------------------------------------------------------------------

    def __init__(self):
        # initialise start time
        global starttime
        starttime = str(datetime.now().strftime("%d-%m-%Y_%H-%M-%S"))

        # get the root directory of project
        strpath = 'C:/Users/ANSUMAN/Documents/PyCharm/Project/'
        # strpath = os.path.realpath(os.getcwd())

        # get the result file path
        global resultpath
        resultpath = os.path.abspath(strpath + '/Result/')

        # get the datasheet file path
        global datapath
        datapath = os.path.abspath(strpath + '/Test/')

        # get the Function Library file path
        global libpath
        libpath = os.path.abspath(strpath + '/Library/')

        # initialise driver
        global driver
        driver = webdriver.Chrome()

        # read page object model
        global pomdict
        try:
            pomdict
        except NameError:
            elmtdict = {}
            wslist = []
            pomsheet = libpath + '/appdetails.xlsx'
            wb = load_workbook(pomsheet)
            # get all sheets
            for wsname in wb.sheetnames:
                if wsname != 'AppDetails':
                    wslist.append(wb[wsname])
            # create dictionary from all the sheets
            for ws in wslist:
                rowcount = len(ws["A"])
                for i in range(2, rowcount + 1):
                    key = str(ws.title) + '.' + str(ws.cell(row=i, column=1).value)
                    val = str(ws.cell(row=i, column=2).value) + '|' + str(ws.cell(row=i, column=3).value)
                    elmtdict[key] = val
                    if key == 'None':
                        break
            pomdict = elmtdict

        # initialise the logger
        global logwriter
        try:
            logwriter
        except NameError:
            with open(resultpath + '/ExecTrace.log', 'w'):
                pass
            log_handler = FileHandler(resultpath + '/ExecTrace.log')
            log_handler.push_application()
            log = Logger('Exec_Status')
            logwriter = log

        # initialise the report
        global reportwriter
        try:
            reportwriter
        except NameError:
            report_handler = FPDF('P', 'mm', 'A4')
            report_handler.add_page()
            reportwriter = report_handler

        # load data from yaml file
        global datahandler
        file = open(datapath + '/data.yml', 'r')
        yml = YAML()
        datadict = yml.load(file)
        datahandler = datadict

    def endrun(self):
        driver.close()
        driver.quit()
        reportwriter.output(resultpath + '/' + starttime + '_ExecReport.pdf', 'F')
        # zip_handler is zipfile handler
        zip_handler = zipfile.ZipFile(resultpath + '/archived/Result_' + starttime + '.zip', 'w')
        for file in os.listdir(resultpath):
            if os.path.isfile(os.path.join(resultpath, file)):
                zip_handler.write(os.path.abspath(os.path.join(resultpath, file)),
                                  os.path.basename(os.path.join(resultpath, file)))
        zip_handler.close()



